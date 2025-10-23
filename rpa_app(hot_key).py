import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

import numpy as np
import pyautogui
import time
import xlrd
import pyperclip
import os
import sys
import threading
from openpyxl import Workbook
from openpyxl.styles import Font
import tempfile
import keyboard  # 新增：用于全局热键监听


# 获取资源的绝对路径（用于打包）
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class RPAApp:
    def __init__(self, root):
        self.root = root
        self.root.title("RPA 自动化工具")
        self.root.geometry("800x850")  # 增加窗口高度以容纳新控件
        self.root.resizable(True, True)

        # 控制变量
        self.is_running = False
        self.current_sheet = None
        self.excel_dir = None  # 存储Excel文件所在目录
        self.hotkey_enabled = False  # 热键启用状态
        self.stop_hotkey = "ctrl+shift+q"  # 默认停止热键
        self.interval_time = 0.01  # 默认时间间隔

        # 创建界面
        self.create_widgets()

    def create_widgets(self):
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(3, weight=1)

        # 标题
        title_label = ttk.Label(main_frame, text="RPA 自动化工具", font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 15))

        # Excel 文件选择
        ttk.Label(main_frame, text="Excel 脚本文件:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.file_path = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.file_path, width=50).grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5,
                                                                          padx=(5, 5))
        ttk.Button(main_frame, text="浏览", command=self.browse_file).grid(row=1, column=2, pady=5)

        # 执行模式选择
        mode_frame = ttk.Frame(main_frame)
        mode_frame.grid(row=2, column=0, columnspan=3, pady=10, sticky=tk.W)

        ttk.Label(mode_frame, text="执行模式:").pack(side=tk.LEFT)
        self.execution_mode = tk.StringVar(value="1")
        ttk.Radiobutton(mode_frame, text="执行一次", variable=self.execution_mode, value="1").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(mode_frame, text="循环执行", variable=self.execution_mode, value="2").pack(side=tk.LEFT, padx=5)

        # 循环次数输入
        loop_frame = ttk.Frame(main_frame)
        loop_frame.grid(row=3, column=0, columnspan=3, pady=5, sticky=tk.W)

        ttk.Label(loop_frame, text="循环次数 (0=无限循环):").pack(side=tk.LEFT)
        self.loop_count = tk.StringVar(value="0")
        self.loop_entry = ttk.Entry(loop_frame, textvariable=self.loop_count, width=10)
        self.loop_entry.pack(side=tk.LEFT, padx=5)

        # 时间间隔设置
        interval_frame = ttk.Frame(main_frame)
        interval_frame.grid(row=4, column=0, columnspan=3, pady=10, sticky=tk.W)

        ttk.Label(interval_frame, text="操作间隔时间(秒):").pack(side=tk.LEFT)
        self.interval_var = tk.StringVar(value="0.01")
        interval_entry = ttk.Entry(interval_frame, textvariable=self.interval_var, width=10)
        interval_entry.pack(side=tk.LEFT, padx=5)
        ttk.Button(interval_frame, text="设置间隔", command=self.set_interval).pack(side=tk.LEFT, padx=5)

        # 间隔说明
        interval_help = ttk.Label(interval_frame, text="(默认0.01秒)", foreground="gray")
        interval_help.pack(side=tk.LEFT, padx=5)

        # 停止热键设置
        hotkey_frame = ttk.Frame(main_frame)
        hotkey_frame.grid(row=5, column=0, columnspan=3, pady=10, sticky=tk.W)

        ttk.Label(hotkey_frame, text="停止热键:").pack(side=tk.LEFT)
        self.hotkey_var = tk.StringVar(value="ctrl+shift+q")
        hotkey_entry = ttk.Entry(hotkey_frame, textvariable=self.hotkey_var, width=15)
        hotkey_entry.pack(side=tk.LEFT, padx=5)
        ttk.Button(hotkey_frame, text="设置热键", command=self.set_hotkey).pack(side=tk.LEFT, padx=5)

        # 热键说明
        hotkey_help = ttk.Label(hotkey_frame, text="(例如: ctrl+shift+q, alt+s, f12)", foreground="gray")
        hotkey_help.pack(side=tk.LEFT, padx=5)

        # 操作按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=6, column=0, columnspan=3, pady=10)

        ttk.Button(button_frame, text="创建示例 Excel 文件", command=self.create_example_excel).pack(side=tk.LEFT,
                                                                                                     padx=5)
        ttk.Button(button_frame, text="开始执行", command=self.start_execution).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="停止执行", command=self.stop_execution).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="获取坐标", command=self.get_coordinates).pack(side=tk.LEFT, padx=5)

        # 日志输出
        ttk.Label(main_frame, text="执行日志:").grid(row=7, column=0, sticky=tk.W, pady=(10, 5))
        self.log_text = scrolledtext.ScrolledText(main_frame, width=80, height=20, state=tk.DISABLED)
        self.log_text.grid(row=8, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)

    def set_interval(self):
        """设置操作间隔时间"""
        try:
            interval = float(self.interval_var.get().strip())
            if interval < 0:
                messagebox.showerror("错误", "间隔时间不能为负数")
                return
            self.interval_time = interval
            self.log(f"操作间隔时间已设置为: {interval} 秒")
        except ValueError:
            messagebox.showerror("错误", "请输入有效的数字")

    def set_hotkey(self):
        """设置停止热键"""
        new_hotkey = self.hotkey_var.get().strip().lower()

        if not new_hotkey:
            messagebox.showerror("错误", "请输入热键组合")
            return

        # 验证热键格式（简单验证）
        valid_modifiers = ['ctrl', 'shift', 'alt', 'win']
        valid_single_keys = ['f1', 'f2', 'f3', 'f4', 'f5', 'f6', 'f7', 'f8', 'f9', 'f10', 'f11', 'f12',
                             'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm',
                             'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z',
                             '0', '1', '2', '3', '4', '5', '6', '7', '8', '9']

        keys = new_hotkey.split('+')
        if len(keys) == 1:
            # 单键
            if keys[0] not in valid_single_keys:
                messagebox.showerror("错误", f"无效的热键: {new_hotkey}\n请使用有效的按键组合")
                return
        else:
            # 组合键
            modifiers = keys[:-1]
            main_key = keys[-1]

            for mod in modifiers:
                if mod not in valid_modifiers:
                    messagebox.showerror("错误", f"无效的修饰键: {mod}\n有效的修饰键: ctrl, shift, alt, win")
                    return

            if main_key not in valid_single_keys:
                messagebox.showerror("错误", f"无效的主按键: {main_key}")
                return

        # 移除旧的热键（如果已设置）
        if self.hotkey_enabled:
            try:
                keyboard.remove_hotkey(self.stop_hotkey)
            except:
                pass

        # 设置新的热键
        self.stop_hotkey = new_hotkey
        self.log(f"停止热键已设置为: {self.stop_hotkey}")
        messagebox.showinfo("成功", f"停止热键已设置为: {self.stop_hotkey}\n\n在循环执行过程中按下此热键可停止程序")

    def register_hotkey(self):
        """注册全局热键"""
        try:
            # 先尝试移除可能存在的旧热键
            try:
                keyboard.remove_hotkey(self.stop_hotkey)
            except:
                pass

            # 注册新热键
            keyboard.add_hotkey(self.stop_hotkey, self.hotkey_stop)
            self.hotkey_enabled = True
            self.log(f"全局停止热键已启用: {self.stop_hotkey}")
        except Exception as e:
            self.log(f"注册热键失败: {str(e)}")
            self.hotkey_enabled = False

    def unregister_hotkey(self):
        """取消注册全局热键"""
        try:
            keyboard.remove_hotkey(self.stop_hotkey)
            self.hotkey_enabled = False
            self.log("全局停止热键已禁用")
        except:
            pass

    def hotkey_stop(self):
        """热键停止回调函数"""
        if self.is_running:
            self.log(f"检测到停止热键 {self.stop_hotkey} 被按下，正在停止...")
            self.stop_execution()

    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="选择 Excel 脚本文件",
            filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.file_path.set(filename)
            # 保存Excel文件所在目录
            self.excel_dir = os.path.dirname(filename)

    def get_coordinates(self):
        """获取当前鼠标坐标"""
        messagebox.showinfo("获取坐标", "请在5秒内将鼠标移动到目标位置...")
        self.root.iconify()  # 最小化窗口

        # 等待5秒让用户移动鼠标
        time.sleep(5)

        # 获取当前鼠标位置
        x, y = pyautogui.position()

        # 恢复窗口
        self.root.deiconify()

        # 显示坐标
        messagebox.showinfo("坐标获取成功", f"当前鼠标坐标: ({x}, {y})\n\n坐标已复制到剪贴板")

        # 复制坐标到剪贴板
        pyperclip.copy(f"{x},{y}")

        # 记录到日志
        self.log(f"获取坐标: ({x}, {y}) - 已复制到剪贴板")

    def create_example_excel(self):
        try:
            # 创建示例 Excel 文件
            wb = Workbook()
            ws = wb.active
            ws.title = "自动化脚本"

            # 添加表头
            headers = ["操作类型", "内容", "重试次数", "说明"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)

            # 添加示例数据
            examples = [
                [1, "button.png", 1, "单击左键 - 请将button.png图片文件放在Excel文件同目录下"],
                [5, 2, 0, "等待2秒"],
                [4, "Hello World", 0, "输入文本"],
                [6, -100, 0, "向下滚动"],
                [1, "login.png", 3, "单击登录按钮，最多重试3次 - 请将login.png放在Excel文件同目录下"],
                [7, "500,300", 1, "单击坐标(500,300) - 使用获取坐标功能获取坐标"]
            ]

            for row, example in enumerate(examples, 2):
                for col, value in enumerate(example, 1):
                    ws.cell(row=row, column=col, value=value)

            # 保存文件
            filename = filedialog.asksaveasfilename(
                title="保存示例 Excel 文件",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )

            if filename:
                wb.save(filename)
                self.log("示例 Excel 文件已创建: " + filename)
                messagebox.showinfo("成功",
                                    "示例 Excel 文件已创建成功！\n\n注意：请将图片文件(.png)放在Excel文件同目录下\n\n新增功能：可以使用'获取坐标'按钮获取鼠标坐标")

        except Exception as e:
            self.log("创建示例 Excel 文件时出错: " + str(e))
            messagebox.showerror("错误", "创建示例 Excel 文件时出错: " + str(e))

    def start_execution(self):
        if not self.file_path.get():
            messagebox.showerror("错误", "请先选择 Excel 脚本文件")
            return

        if not os.path.exists(self.file_path.get()):
            messagebox.showerror("错误", "选择的文件不存在")
            return

        if self.is_running:
            messagebox.showwarning("警告", "程序已在运行中")
            return

        # 验证循环次数输入
        loop_count_str = self.loop_count.get().strip()
        if not loop_count_str:
            messagebox.showerror("错误", "请输入循环次数 (0表示无限循环)")
            return

        try:
            loop_count = int(loop_count_str)
            if loop_count < 0:
                messagebox.showerror("错误", "循环次数不能为负数")
                return
        except ValueError:
            messagebox.showerror("错误", "循环次数必须是数字")
            return

        # 注册全局热键
        self.register_hotkey()

        self.is_running = True
        self.log("开始执行自动化脚本...")
        self.log(f"停止热键: {self.stop_hotkey} (在循环执行过程中按下可停止程序)")
        self.log(f"操作间隔时间: {self.interval_time} 秒")

        # 在新线程中执行自动化任务
        thread = threading.Thread(target=self.execute_automation)
        thread.daemon = True
        thread.start()

    def stop_execution(self):
        self.is_running = False
        # 取消注册热键
        self.unregister_hotkey()
        self.log("停止执行命令已发送...")

    def execute_automation(self):
        try:
            # 打开 Excel 文件
            wb = xlrd.open_workbook(self.file_path.get())
            sheet = wb.sheet_by_index(0)

            # 数据检查
            if not self.data_check(sheet):
                self.log("数据检查未通过，请检查 Excel 文件内容!")
                self.unregister_hotkey()  # 取消热键注册
                return

            self.log("数据检查通过，开始执行脚本...")

            # 获取循环次数
            loop_count = int(self.loop_count.get())

            # 根据执行模式执行
            if self.execution_mode.get() == "1":
                # 执行一次
                self.main_work(sheet)
                self.log("执行完成!")
            else:
                # 循环执行
                if loop_count == 0:
                    # 无限循环
                    count = 1
                    while self.is_running:
                        self.log(f"第 {count} 次循环执行...")
                        self.main_work(sheet)
                        time.sleep(self.interval_time)
                        count += 1
                else:
                    # 有限循环
                    for i in range(loop_count):
                        if not self.is_running:
                            break
                        self.log(f"第 {i + 1} 次循环执行...")
                        self.main_work(sheet)
                        time.sleep(self.interval_time)

            self.log("自动化任务执行完毕")

        except Exception as e:
            self.log(f"执行过程中出错: {str(e)}")
        finally:
            self.is_running = False
            self.unregister_hotkey()  # 确保热键被取消注册

    def data_check(self, sheet):
        try:
            if sheet.nrows < 2:
                self.log("Excel 文件中没有数据")
                return False

            for i in range(1, sheet.nrows):
                cmd_type = sheet.row(i)[0]
                if cmd_type.ctype != 2 or cmd_type.value not in [1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0]:
                    self.log(f'第 {i + 1} 行, 第1列数据有误')
                    return False

                cmd_value = sheet.row(i)[1]
                if cmd_type.value in [1.0, 2.0, 3.0]:
                    if cmd_value.ctype != 1:
                        self.log(f'第 {i + 1} 行, 第2列数据有误')
                        return False
                elif cmd_type.value == 4.0:
                    if cmd_value.ctype == 0:
                        self.log(f'第 {i + 1} 行, 第2列数据有误')
                        return False
                elif cmd_type.value == 5.0:
                    if cmd_value.ctype != 2:
                        self.log(f'第 {i + 1} 行, 第2列数据有误')
                        return False
                elif cmd_type.value == 6.0:
                    if cmd_value.ctype != 2:
                        self.log(f'第 {i + 1} 行, 第2列数据有误')
                        return False
                elif cmd_type.value == 7.0:  # 坐标点击
                    if cmd_value.ctype != 1:
                        self.log(f'第 {i + 1} 行, 第2列数据有误')
                        return False
                    # 检查坐标格式
                    try:
                        coords = cmd_value.value.split(';')
                        if len(coords) != 2:
                            self.log(f'第 {i + 1} 行, 坐标格式错误，应为"x;y"')
                            return False
                        x, y = int(coords[0]), int(coords[1])
                    except ValueError:
                        self.log(f'第 {i + 1} 行, 坐标格式错误，应为"x;y"')
                        return False

            return True
        except Exception as e:
            self.log(f"数据检查出错: {str(e)}")
            return False

    def main_work(self, sheet):
        for i in range(1, sheet.nrows):
            if not self.is_running:
                break

            try:
                cmd_type = sheet.row(i)[0]
                if cmd_type.value == 1.0:
                    img = sheet.row(i)[1].value
                    reTry = 1
                    if sheet.row(i)[2].ctype == 2 and sheet.row(i)[2].value != 0:
                        reTry = sheet.row(i)[2].value
                    success = self.mouse_click(1, "left", img, reTry)
                    if success:
                        self.log(f"单击左键: {img}")
                    else:
                        self.log(f"跳过第 {i + 1} 行命令：单击左键 {img}")

                elif cmd_type.value == 2.0:
                    img = sheet.row(i)[1].value
                    reTry = 1
                    if sheet.row(i)[2].ctype == 2 and sheet.row(i)[2].value != 0:
                        reTry = sheet.row(i)[2].value
                    success = self.mouse_click(2, "left", img, reTry)
                    if success:
                        self.log(f"双击左键: {img}")
                    else:
                        self.log(f"跳过第 {i + 1} 行命令：双击左键 {img}")

                elif cmd_type.value == 3.0:
                    img = sheet.row(i)[1].value
                    reTry = 1
                    if sheet.row(i)[2].ctype == 2 and sheet.row(i)[2].value != 0:
                        reTry = sheet.row(i)[2].value
                    success = self.mouse_click(1, "right", img, reTry)
                    if success:
                        self.log(f"右键: {img}")
                    else:
                        self.log(f"跳过第 {i + 1} 行命令：右键 {img}")

                elif cmd_type.value == 4.0:
                    input_value = sheet.row(i)[1].value
                    pyperclip.copy(input_value)
                    pyautogui.hotkey('ctrl', 'v')
                    time.sleep(0.5)
                    self.log(f"输入: {input_value}")

                elif cmd_type.value == 5.0:
                    wait_time = sheet.row(i)[1].value
                    time.sleep(wait_time)
                    self.log(f"等待 {wait_time} 秒")

                elif cmd_type.value == 6.0:
                    scroll = sheet.row(i)[1].value
                    pyautogui.scroll(int(scroll))
                    self.log(f"滚轮滑动 {int(scroll)} 距离")

                elif cmd_type.value == 7.0:  # 坐标点击
                    coord_str = sheet.row(i)[1].value
                    reTry = 1
                    if sheet.row(i)[2].ctype == 2 and sheet.row(i)[2].value != 0:
                        reTry = sheet.row(i)[2].value
                    success = self.coordinate_click(coord_str, reTry)
                    if success:
                        self.log(f"坐标点击: {coord_str}")
                    else:
                        self.log(f"跳过第 {i + 1} 行命令：坐标点击 {coord_str}")

            except Exception as e:
                self.log(f"执行第 {i + 1} 行命令时发生错误: {str(e)}，跳过此命令")

    def mouse_click(self, click_times, l_or_r, img, retry):
        # 解析图片路径
        img_path = self.resolve_image_path(img)

        # 检查图片文件是否存在
        if not os.path.exists(img_path):
            self.log(f"错误：图片文件 '{img}' 不存在（尝试路径: {img_path}）")
            return False

        self.log(f"正在查找图片: {img_path}")

        if retry == 1:
            attempt = 0
            while attempt < 3:  # 最多尝试3次
                try:
                    location = pyautogui.locateCenterOnScreen(img_path, confidence=0.8)
                    if location is not None:
                        self.log(f"找到图片，位置: {location}")
                        pyautogui.click(location.x, location.y, clicks=click_times,
                                        interval=0.2, duration=0.2, button=l_or_r)
                        return True
                except pyautogui.ImageNotFoundException:
                    pass

                attempt += 1
                time.sleep(self.interval_time)

            self.log(f"未找到匹配图片 '{img}'，跳过此命令")
            return False

        elif retry == -1:
            while self.is_running:
                try:
                    location = pyautogui.locateCenterOnScreen(img_path, confidence=0.8)
                    if location is not None:
                        pyautogui.click(location.x, location.y, clicks=click_times,
                                        interval=0.2, duration=0.2, button=l_or_r)
                except pyautogui.ImageNotFoundException:
                    pass
                time.sleep(self.interval_time)
            return True

        elif retry > 1:
            i = 1
            while i < retry + 1 and self.is_running:
                try:
                    i += 1
                    location = pyautogui.locateCenterOnScreen(img_path, confidence=0.8)
                    if location is not None:
                        pyautogui.click(location.x, location.y, clicks=click_times,
                                        interval=0.2, duration=0.2, button=l_or_r)
                        self.log(f"重复执行第 {i-1} 次")
                except pyautogui.ImageNotFoundException:
                    self.log(f"第 {i-1} 次尝试未找到图片，继续重试")
                time.sleep(self.interval_time)
            return True
        return True

    def coordinate_click(self, coord_str, retry):
        """根据坐标进行点击"""
        try:
            # 解析坐标
            coords = coord_str.split(';')
            if len(coords) != 2:
                self.log(f"坐标格式错误: {coord_str}")
                return False

            x, y = int(coords[0]), int(coords[1])

            # 执行点击
            if retry == 1:
                pyautogui.click(x, y)
                return True
            elif retry > 1:
                i = 1
                while i <= retry and self.is_running:
                    i += 1
                    pyautogui.click(x, y)
                    self.log(f"坐标点击重复执行第 {i-1} 次")
                    time.sleep(self.interval_time)
                return True
            elif retry == -1:
                while self.is_running:
                    pyautogui.click(x, y)
                    time.sleep(self.interval_time)
                return True

            return True
        except Exception as e:
            self.log(f"坐标点击出错: {str(e)}")
            return False

    def resolve_image_path(self, img):
        """
        解析图片路径，按以下顺序尝试：
        1. 如果已经是绝对路径，直接使用
        2. 尝试在Excel文件同目录下查找
        3. 尝试在EXE文件同目录下查找
        4. 尝试在当前工作目录下查找
        """
        # 如果是绝对路径，直接返回
        if os.path.isabs(img):
            return img

        # 尝试在Excel文件同目录下查找
        if self.excel_dir:
            excel_dir_path = os.path.join(self.excel_dir, img)
            if os.path.exists(excel_dir_path):
                return excel_dir_path

        # 尝试在EXE文件同目录下查找
        if getattr(sys, 'frozen', False):
            # 如果是打包后的EXE
            exe_dir = os.path.dirname(sys.executable)
            exe_dir_path = os.path.join(exe_dir, img)
            if os.path.exists(exe_dir_path):
                return exe_dir_path
        else:
            # 如果是Python脚本
            script_dir = os.path.dirname(os.path.abspath(__file__))
            script_dir_path = os.path.join(script_dir, img)
            if os.path.exists(script_dir_path):
                return script_dir_path

        # 尝试在当前工作目录下查找
        cwd_path = os.path.join(os.getcwd(), img)
        if os.path.exists(cwd_path):
            return cwd_path

        # 如果都找不到，返回原始路径（可能会在后续检查中失败）
        return img

    def log(self, message):
        # 在日志文本框中添加消息
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"{time.strftime('%H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        # 强制更新界面
        self.root.update_idletasks()


def main():
    root = tk.Tk()
    app = RPAApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()